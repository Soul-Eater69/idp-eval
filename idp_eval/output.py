"""Evaluation result publishing: normalized records and output writers.

Evaluation happens once and produces :class:`~idp_eval.models.EvaluationResult`
objects. Those are normalized into :class:`EvaluationRecord` rows and handed to
one or more :class:`EvaluationWriter` implementations. The canonical Phoenix
persistence mechanism is **native span annotations** (not span attributes);
Excel is an independent flat export. Writers never trigger re-evaluation.
"""

from __future__ import annotations

import json
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Protocol, runtime_checkable

from idp_eval.models import EvaluationResult
from idp_eval.rendering import render_value

# Phoenix's accepted annotator kinds.
ANNOTATOR_KINDS = ("LLM", "CODE", "HUMAN")


def validate_annotator_kind(kind: str) -> str:
    """Returns ``kind`` if it is a supported annotator kind, else raises."""
    if kind not in ANNOTATOR_KINDS:
        raise ValueError(
            f"Unsupported annotator_kind {kind!r}; expected one of "
            f"{ANNOTATOR_KINDS}."
        )
    return kind


class PersistenceError(RuntimeError):
    """Raised when writing results fails after evaluation already succeeded.

    The computed results are preserved on :attr:`results` so callers never lose
    them just because persistence failed.
    """

    def __init__(self, message: str, results: dict[str, EvaluationResult]):
        super().__init__(message)
        self.results = results


@dataclass(frozen=True)
class EvaluationRecord:
    """One normalized (case + metric result) row for publishing.

    ``input`` is the optional rendered task/query used for summary reporting.
    ``span_id`` targets the root ``idp_eval.evaluate`` span for native Phoenix
    annotations; it is required for Phoenix output and ``None`` when tracing is
    inactive. ``identifier`` is an optional upsert key for the Phoenix annotation
    API and is never required.
    """

    metric: str
    score: float | None
    label: str | None
    explanation: str | None
    annotator_kind: str
    status: str = "success"
    case_id: str | None = None
    run_name: str | None = None
    dataset_name: str | None = None
    case_fingerprint: str | None = None
    evaluation_fingerprint: str | None = None
    context: str | None = None
    output: str | None = None
    instructions: str | None = None
    evaluation_scope: str | None = None
    retrieved_documents_json: str | None = None
    retrieved_documents: str | None = None
    metadata: dict[str, Any] | None = None
    trace_id: str | None = None
    span_id: str | None = None
    identifier: str | None = None
    details: dict[str, Any] | None = None
    timestamp: str = field(
        default_factory=lambda: datetime.now(timezone.utc).isoformat()
    )
    input: str | None = None

    @classmethod
    def from_result(
        cls,
        result: EvaluationResult,
        *,
        annotator_kind: str,
        status: str = "success",
        case_id: str | None = None,
        input: str | None = None,
        run_name: str | None = None,
        dataset_name: str | None = None,
        case_fingerprint: str | None = None,
        evaluation_fingerprint: str | None = None,
        context: str | None = None,
        output: str | None = None,
        instructions: str | None = None,
        evaluation_scope: str | None = None,
        retrieved_documents_json: str | None = None,
        retrieved_documents: str | None = None,
        metadata: dict[str, Any] | None = None,
        trace_id: str | None = None,
        span_id: str | None = None,
        identifier: str | None = None,
    ) -> "EvaluationRecord":
        """Builds a record from an :class:`EvaluationResult`."""
        return cls(
            metric=result.metric,
            score=result.score,
            label=result.label,
            explanation=result.explanation,
            annotator_kind=validate_annotator_kind(annotator_kind),
            status=status,
            case_id=case_id,
            input=input,
            run_name=run_name,
            dataset_name=dataset_name,
            case_fingerprint=case_fingerprint,
            evaluation_fingerprint=evaluation_fingerprint,
            context=context,
            output=output,
            instructions=instructions,
            evaluation_scope=evaluation_scope,
            retrieved_documents_json=retrieved_documents_json,
            retrieved_documents=retrieved_documents,
            metadata=metadata,
            trace_id=trace_id,
            span_id=span_id,
            identifier=identifier,
            details=result.details,
        )


@runtime_checkable
class EvaluationWriter(Protocol):
    """Destination for normalized evaluation records."""

    def write(self, records: list[EvaluationRecord]) -> None:
        """Persists a batch of records."""


@runtime_checkable
class EvaluationCheckpoint(Protocol):
    """Checkpoint lookup implemented by resumable result destinations."""

    def load_successful_result(
        self,
        *,
        run_name: str | None,
        dataset_name: str | None,
        case_id: str | None,
        evaluation_fingerprint: str,
        metric: str,
    ) -> EvaluationResult | None:
        """Returns an exact successful result, or ``None`` when work remains."""


def _safe_metadata(details: dict[str, Any] | None) -> dict[str, Any] | None:
    """Returns compact, JSON-safe annotation metadata (or ``None``).

    Guards against dumping oversized content into annotation metadata by
    round-tripping through JSON with a string fallback.
    """
    if not details:
        return None
    return json.loads(json.dumps(details, default=str))


def _make_span_annotation(payload: dict[str, Any]):
    """Builds a Phoenix ``SpanAnnotationData`` from a normalized payload.

    Imported lazily so Excel-only usage never imports Phoenix. Patchable in
    tests to avoid requiring the Phoenix package.
    """
    from phoenix.client.resources.spans import SpanAnnotationData

    return SpanAnnotationData(**payload)


def _flush_tracer_provider() -> None:
    """Best-effort flush of pending spans so Phoenix can ingest the target span.

    A span annotation can only attach to a span Phoenix has already received, so
    we flush before logging. No-op when OpenTelemetry is unavailable or the
    active provider has no ``force_flush``.
    """
    try:
        from opentelemetry import trace

        flush = getattr(trace.get_tracer_provider(), "force_flush", None)
        if callable(flush):
            flush()
    except Exception:  # noqa: BLE001 - flushing is best-effort
        pass


class PhoenixEvaluationWriter:
    """Writes each result as a native Phoenix span annotation on the case span.

    Annotations for one case are sent in a single batched
    ``client.spans.log_span_annotations`` call. Requires an active trace (a
    ``span_id`` on every record); if Phoenix output is requested without tracing,
    this raises rather than silently claiming persistence.

    Because the framework logs annotations right after the case span closes, the
    target span may not be ingested by Phoenix yet — Phoenix returns 404 in that
    window and drops the annotation. So the writer flushes pending spans, then
    retries the (idempotent-per-span) write on 404 up to ``ingest_timeout``
    seconds before surfacing the error.
    """

    def __init__(
        self,
        client: Any | None = None,
        *,
        base_url: str | None = None,
        api_key: str | None = None,
        ingest_timeout: float = 10.0,
        poll_interval: float = 0.5,
    ):
        """Args:
            client: An injected Phoenix client. When provided it is used as-is
                (``base_url`` / ``api_key`` are ignored), which keeps dependency
                injection for tests intact.
            base_url: Phoenix REST base URL. ``None`` → Phoenix reads
                ``PHOENIX_BASE_URL`` (default ``http://localhost:6006``).
            api_key: API key for authenticated Phoenix. ``None`` → Phoenix reads
                ``PHOENIX_API_KEY``. Sent by the Phoenix client as a bearer token;
                never logged or written into annotations.
            ingest_timeout: Max seconds to wait for the target span to be
                ingested before failing.
            poll_interval: Delay between write retries while waiting for ingest.
        """
        self._client = client
        self._base_url = base_url
        self._api_key = api_key
        self._ingest_timeout = ingest_timeout
        self._poll_interval = poll_interval

    def _get_client(self) -> Any:
        """Returns the injected client, or lazily builds a Phoenix ``Client``.

        With no explicit ``base_url`` / ``api_key``, ``Client()`` is used so
        Phoenix resolves ``PHOENIX_BASE_URL`` / ``PHOENIX_API_KEY`` from the
        environment. Explicit values are forwarded to the native client kwargs;
        no Authorization header is constructed by hand.
        """
        if self._client is None:
            from phoenix.client import Client

            kwargs: dict[str, Any] = {}
            if self._base_url is not None:
                kwargs["base_url"] = self._base_url
            if self._api_key is not None:
                kwargs["api_key"] = self._api_key
            self._client = Client(**kwargs)
        return self._client

    @staticmethod
    def _payload(record: EvaluationRecord) -> dict[str, Any]:
        """Maps a record to Phoenix ``SpanAnnotationData`` kwargs.

        ``score`` / ``label`` / ``explanation`` are omitted when ``None`` (a
        not-applicable result is never coerced into a fake ``0.0``).
        """
        result: dict[str, Any] = {}
        if record.score is not None:
            result["score"] = record.score
        if record.label is not None:
            result["label"] = record.label
        if record.explanation is not None:
            result["explanation"] = record.explanation

        payload: dict[str, Any] = {
            "name": record.metric,
            "span_id": record.span_id,
            "annotator_kind": record.annotator_kind,
            "result": result,
        }
        if record.identifier is not None:
            payload["identifier"] = record.identifier
        metadata = _safe_metadata(record.details)
        if metadata is not None:
            payload["metadata"] = metadata
        return payload

    def write(self, records: list[EvaluationRecord]) -> None:
        if not records:
            return
        if any(record.span_id is None for record in records):
            raise RuntimeError(
                "Phoenix output requires an active trace: no target span id. "
                "Call register_tracing() before evaluating."
            )
        annotations = [_make_span_annotation(self._payload(r)) for r in records]
        _flush_tracer_provider()
        self._log_with_retry(annotations)

    def _log_with_retry(self, annotations: list[Any]) -> None:
        """Logs the batch, retrying on 404 until the target span is ingested."""
        import time

        import httpx

        client = self._get_client()
        deadline = time.monotonic() + self._ingest_timeout
        while True:
            try:
                # sync=True validates persistence (and raises 404 if the target
                # span is not ingested yet); the async default would silently
                # drop the annotation.
                client.spans.log_span_annotations(
                    span_annotations=annotations, sync=True
                )
                return
            except httpx.HTTPStatusError as exc:
                # 404 => target span not yet ingested by Phoenix; retry briefly.
                if (
                    exc.response.status_code != 404
                    or time.monotonic() >= deadline
                ):
                    raise
                time.sleep(self._poll_interval)


DEFAULT_REPORT_FIELDS = ("input", "context", "output", "instructions")
_CASE_REPORT_FIELDS = frozenset(
    {"input", "context", "output", "instructions", "retrieved_documents"}
)
_SUMMARY_PREFIX = ("run_name", "dataset_name", "key_id")
_SUMMARY_SUFFIX = (
    "metric",
    "status",
    "score",
    "label",
    "explanation",
    "timestamp",
    "raw_details_json",
)
_RESERVED_REPORT_COLUMNS = frozenset((*_SUMMARY_PREFIX, *_SUMMARY_SUFFIX))
_CHECKPOINT_SHEET = "_idp_eval_checkpoint"
_CHECKPOINT_COLUMNS = (
    "run_name",
    "dataset_name",
    "case_id",
    "case_fingerprint",
    "evaluation_fingerprint",
    "metric",
    "status",
    "trace_id",
    "span_id",
    "annotator_kind",
    "timestamp",
    "score",
    "label",
    "explanation",
    "raw_details_json",
    "input",
    "context",
    "output",
    "instructions",
    "evaluation_scope",
    "retrieved_documents_json",
    "retrieved_documents",
    "metadata_json",
    "identifier",
)


def validate_report_fields(
    report_fields: list[str] | tuple[str, ...] | None,
) -> tuple[str, ...]:
    """Validates and preserves ordered visible Excel case fields."""
    if not isinstance(report_fields, (list, tuple, type(None))):
        raise ValueError("report_fields must be a list, tuple, or None.")
    fields = DEFAULT_REPORT_FIELDS if report_fields is None else tuple(report_fields)

    seen_fields: set[str] = set()
    seen_columns: set[str] = set(_RESERVED_REPORT_COLUMNS)
    for field_name in fields:
        if not isinstance(field_name, str):
            raise ValueError("report_fields entries must be strings.")
        if field_name in seen_fields:
            raise ValueError(f"Duplicate report field: {field_name!r}.")
        seen_fields.add(field_name)

        if field_name in _CASE_REPORT_FIELDS:
            column = field_name
        elif field_name.startswith("metadata."):
            column = field_name.removeprefix("metadata.")
            if not column or "." in column:
                raise ValueError(
                    "Metadata report fields must use one non-empty key, for "
                    "example 'metadata.theme_id'."
                )
        else:
            supported = ", ".join(sorted(_CASE_REPORT_FIELDS))
            raise ValueError(
                f"Unknown report field {field_name!r}; supported fields are "
                f"{supported}, plus metadata.<key>."
            )
        if column in seen_columns:
            raise ValueError(
                f"Report field {field_name!r} conflicts with visible Excel "
                f"column {column!r}."
            )
        seen_columns.add(column)
    return fields


def _report_column(field_name: str) -> str:
    return field_name.removeprefix("metadata.")


def _summary_columns(report_fields: tuple[str, ...]) -> tuple[str, ...]:
    return (*_SUMMARY_PREFIX, *map(_report_column, report_fields), *_SUMMARY_SUFFIX)

# Identity columns prefixed onto every metric-detail sheet row.
_ITEM_IDENTITY_COLUMNS = (
    "run_name",
    "dataset_name",
    "key_id",
    "metric",
)


@dataclass(frozen=True)
class _ItemSheet:
    """Declarative mapping from a metric's ``details`` to a flat detail sheet.

    ``columns`` pairs each output column name with the field to read from each
    item dict, so one generic loop can flatten any registered metric — no
    per-metric branching in the writer.
    """

    sheet: str
    list_key: str                          # details key holding the item list
    columns: tuple[tuple[str, str], ...]   # (output column, item field)


# Small declarative registry (not per-metric if/else): each entry flattens one
# metric's item list into a dedicated sheet. Arbitrary custom metrics appear
# only in the summary sheet unless they define a repository-supported layout.
_ITEM_SHEETS: dict[str, _ItemSheet] = {
    # Whole-source coverage. ``details["items"]`` is only present in verbose mode
    # (compact results keep details lean), so this sheet fills for verbose runs.
    "coverage": _ItemSheet(
        "coverage_items",
        "items",
        (
            ("item_id", "id"),
            ("source_item", "source_item"),
            ("meaningfully_present", "meaningfully_present"),
            ("fully_present", "fully_present"),
            ("status", "status"),
            ("item_score", "item_score"),
            ("reason", "reason"),
        ),
    ),
    "instruction_adherence": _ItemSheet(
        "instruction_adherence_items",
        "instructions",
        (
            ("instruction_id", "id"),
            ("instruction", "instruction"),
            ("status", "status"),
            ("item_score", "score"),
            ("reason", "reason"),
        ),
    ),
    "faithfulness": _ItemSheet(
        "faithfulness_items",
        "claims",
        (
            ("claim_id", "id"),
            ("claim", "claim"),
            ("status", "status"),
            ("item_score", "item_score"),
            ("reason", "reason"),
        ),
    ),
    "contextual_relevancy": _ItemSheet(
        "contextual_relevancy_items",
        "items",
        (
            ("document_rank", "document_rank"),
            ("context_item", "context_item"),
            ("relevant", "relevant"),
            ("reason", "reason"),
        ),
    ),
    "contextual_recall": _ItemSheet(
        "contextual_recall_items",
        "items",
        (
            ("reference_item", "reference_item"),
            ("captured", "captured"),
            ("reason", "reason"),
        ),
    ),
}

# Retrieval metrics reuse the SAME batched relevance judgments, so their
# documents are written to one shared sheet once per case rather than duplicated
# per metric. They are detected structurally by the ``documents`` detail list.
_RETRIEVAL_DOCUMENTS_SHEET = "retrieval_documents"
_RETRIEVAL_IDENTITY_COLUMNS = (
    "run_name",
    "dataset_name",
    "key_id",
)
_RETRIEVAL_DOCUMENT_COLUMNS = (
    ("rank", "rank"),
    ("document_id", "document_id"),
    ("relevant", "relevant"),
    ("relevance_score", "relevance_score"),
    ("reason", "reason"),
    ("retrieval_score", "retrieval_score"),
)

# Columns given extra width because they hold free text.
_WIDE_COLUMNS = frozenset(
    {
        "explanation",
        "reason",
        "requirement",
        "source_item",
        "instruction",
        "claim",
        "context_item",
        "reference_item",
        "input",
        "context",
        "output",
        "instructions",
        "retrieved_documents",
        "raw_details_json",
    }
)


class ExcelEvaluationWriter:
    """Writes evaluation results to a multi-sheet ``.xlsx`` workbook.

    - ``evaluations``: one row per (case + metric) — the human-readable summary.
    - ``_idp_eval_checkpoint``: hidden technical identity and serialized results.
    - ``<metric>_items``: one row per structured detail item for metrics with a
      registered item layout (coverage requirements / source items, adherence
      instructions), so results are inspectable without reading JSON.

    Each :meth:`write` updates the checkpoint, rebuilds its human-facing
    projections, and saves the workbook. Metrics without a registered detail
    layout appear only in the summary; their full ``details`` remain available
    in ``raw_details_json``. Independent of Phoenix.
    """

    def __init__(
        self,
        path: str,
        *,
        resume: bool = False,
        report_fields: list[str] | tuple[str, ...] | None = None,
    ):
        self._path = path
        self._resume = resume
        self._report_fields = validate_report_fields(report_fields)
        self._summary_columns = _summary_columns(self._report_fields)
        self._workbook: Any | None = None
        self._summary: Any | None = None
        self._checkpoint_sheet: Any | None = None
        self._item_sheets: dict[str, Any] = {}
        self._checkpoint_index: dict[tuple, int] = {}
        if resume:
            self._ensure_workbook()

    def write(self, records: list[EvaluationRecord]) -> None:
        self._ensure_workbook()
        for record in records:
            self._upsert_checkpoint(record)
        self._rebuild_visible_sheets()
        self._workbook.save(self._path)

    def _ensure_workbook(self) -> None:
        """Builds the workbook and summary sheet once, on the first write."""
        if self._workbook is not None:
            return
        from openpyxl import Workbook, load_workbook

        path = Path(self._path)
        if self._resume and path.exists():
            self._workbook = load_workbook(path)
            required_sheets = {"evaluations", _CHECKPOINT_SHEET}
            missing = required_sheets.difference(self._workbook.sheetnames)
            if missing:
                raise ValueError(
                    "Cannot resume Excel workbook: missing required sheet(s) "
                    f"{sorted(missing)} for the resumable schema."
                )
            self._summary = self._workbook["evaluations"]
            self._validate_headers(
                self._summary, self._summary_columns, "evaluations"
            )
            self._checkpoint_sheet = self._workbook[_CHECKPOINT_SHEET]
            self._validate_headers(
                self._checkpoint_sheet, _CHECKPOINT_COLUMNS, _CHECKPOINT_SHEET
            )
            if self._checkpoint_sheet.sheet_state == "visible":
                raise ValueError(
                    "Cannot resume Excel workbook: technical checkpoint sheet "
                    "must be hidden."
                )
            self._rebuild_checkpoint_index()
            return

        self._workbook = Workbook()
        self._summary = self._workbook.active
        self._summary.title = "evaluations"
        self._init_sheet(self._summary, self._summary_columns)
        self._checkpoint_sheet = self._workbook.create_sheet(_CHECKPOINT_SHEET)
        self._init_sheet(self._checkpoint_sheet, _CHECKPOINT_COLUMNS)
        self._checkpoint_sheet.sheet_state = "veryHidden"

    @staticmethod
    def _record_key(record: EvaluationRecord) -> tuple:
        return (
            record.run_name,
            record.dataset_name,
            record.case_id,
            record.evaluation_fingerprint,
        )

    def _checkpoint_values(self, record: EvaluationRecord) -> list[Any]:
        details_json = (
            json.dumps(record.details, default=str)
            if record.details is not None
            else None
        )
        metadata_json = (
            json.dumps(record.metadata, default=str, sort_keys=True)
            if record.metadata is not None
            else None
        )
        return [
            record.run_name,
            record.dataset_name,
            record.case_id,
            record.case_fingerprint,
            record.evaluation_fingerprint,
            record.metric,
            record.status,
            record.trace_id,
            record.span_id,
            record.annotator_kind,
            record.timestamp,
            record.score,
            record.label,
            record.explanation,
            details_json,
            record.input,
            record.context,
            record.output,
            record.instructions,
            record.evaluation_scope,
            record.retrieved_documents_json,
            record.retrieved_documents,
            metadata_json,
            record.identifier,
        ]

    def _upsert_checkpoint(self, record: EvaluationRecord) -> None:
        values = self._checkpoint_values(record)
        if not self._resume:
            self._checkpoint_sheet.append(values)
            return
        key = self._record_key(record)
        row_number = self._checkpoint_index.get(key)
        if row_number is None:
            self._checkpoint_sheet.append(values)
            self._checkpoint_index[key] = self._checkpoint_sheet.max_row
            return
        for column, value in enumerate(values, start=1):
            self._checkpoint_sheet.cell(row=row_number, column=column, value=value)

    def _summary_values(self, record: EvaluationRecord) -> list[Any]:
        selected: list[Any] = []
        for field_name in self._report_fields:
            if field_name.startswith("metadata."):
                key = field_name.removeprefix("metadata.")
                value = (record.metadata or {}).get(key)
                selected.append(self._render_metadata_value(value))
            else:
                selected.append(getattr(record, field_name))
        values = [
            record.run_name,
            record.dataset_name,
            record.case_id,
            *selected,
            record.metric,
            record.status,
            record.score,
            record.label,
            record.explanation,
            record.timestamp,
            (
                json.dumps(record.details, default=str)
                if record.details is not None
                else None
            ),
        ]
        return values

    @staticmethod
    def _render_metadata_value(value: Any) -> str | None:
        if value is None:
            return None
        try:
            return render_value(value) or None
        except (TypeError, ValueError):
            return str(value)

    def load_successful_result(
        self,
        *,
        run_name: str | None,
        dataset_name: str | None,
        case_id: str | None,
        evaluation_fingerprint: str,
        metric: str,
    ) -> EvaluationResult | None:
        """Returns an exact successful checkpoint result, including N/A."""
        self._ensure_workbook()
        row_number = self._checkpoint_index.get(
            (run_name, dataset_name, case_id, evaluation_fingerprint)
        )
        if row_number is None:
            return None
        headers = self._header_index(self._checkpoint_sheet)
        row = self._checkpoint_sheet[row_number]
        if row[headers["status"]].value != "success":
            return None
        persisted_metric = row[headers["metric"]].value
        if persisted_metric != metric:
            raise ValueError(
                "Cannot resume Excel workbook: evaluation fingerprint maps to "
                f"metric {persisted_metric!r}, expected {metric!r}."
            )
        details_value = row[headers["raw_details_json"]].value
        details = None
        if details_value:
            try:
                details = json.loads(details_value)
            except (TypeError, json.JSONDecodeError) as exc:
                raise ValueError(
                    "Cannot resume Excel workbook: invalid raw_details_json for "
                    f"metric {metric!r}."
                ) from exc
            if details is not None and not isinstance(details, dict):
                raise ValueError(
                    "Cannot resume Excel workbook: raw_details_json must contain "
                    "an object or null."
                )
        return EvaluationResult(
            metric=metric,
            score=row[headers["score"]].value,
            label=row[headers["label"]].value,
            explanation=row[headers["explanation"]].value,
            details=details,
        )

    @staticmethod
    def _header_index(sheet: Any) -> dict[str, int]:
        return {cell.value: index for index, cell in enumerate(sheet[1])}

    @classmethod
    def _validate_headers(cls, sheet: Any, expected, name: str) -> None:
        actual = tuple(cell.value for cell in sheet[1])
        if actual != tuple(expected):
            raise ValueError(
                f"Cannot resume Excel workbook: sheet {name!r} predates or "
                "does not match the resumable evaluation schema. Use a new "
                "Excel path or run with resume=False."
            )

    def _rebuild_checkpoint_index(self) -> None:
        headers = self._header_index(self._checkpoint_sheet)
        required = (
            "run_name",
            "dataset_name",
            "case_id",
            "evaluation_fingerprint",
        )
        for row_number in range(2, self._checkpoint_sheet.max_row + 1):
            row = self._checkpoint_sheet[row_number]
            key = tuple(row[headers[name]].value for name in required)
            case_hash = row[headers["case_fingerprint"]].value
            metric = row[headers["metric"]].value
            status = row[headers["status"]].value
            if not key[-1] or not case_hash or not metric:
                raise ValueError(
                    "Cannot resume Excel workbook: an evaluation row is missing "
                    "a required fingerprint or metric identity."
                )
            if status not in ("success", "error"):
                raise ValueError(
                    "Cannot resume Excel workbook: evaluation status must be "
                    f"'success' or 'error', got {status!r}."
                )
            if key in self._checkpoint_index:
                raise ValueError(
                    "Cannot resume Excel workbook: duplicate evaluation identity "
                    f"found for case_id={key[2]!r}."
                )
            self._checkpoint_index[key] = row_number

    def _record_from_checkpoint_row(self, row: Any) -> EvaluationRecord:
        headers = self._header_index(self._checkpoint_sheet)

        def value(name: str) -> Any:
            return row[headers[name]].value

        details = json.loads(value("raw_details_json")) if value(
            "raw_details_json"
        ) else None
        metadata = json.loads(value("metadata_json")) if value(
            "metadata_json"
        ) else None
        return EvaluationRecord(
            metric=value("metric"),
            score=value("score"),
            label=value("label"),
            explanation=value("explanation"),
            annotator_kind=value("annotator_kind"),
            status=value("status"),
            case_id=value("case_id"),
            run_name=value("run_name"),
            dataset_name=value("dataset_name"),
            case_fingerprint=value("case_fingerprint"),
            evaluation_fingerprint=value("evaluation_fingerprint"),
            input=value("input"),
            context=value("context"),
            output=value("output"),
            instructions=value("instructions"),
            evaluation_scope=value("evaluation_scope"),
            retrieved_documents_json=value("retrieved_documents_json"),
            retrieved_documents=value("retrieved_documents"),
            metadata=metadata,
            trace_id=value("trace_id"),
            span_id=value("span_id"),
            identifier=value("identifier"),
            details=details,
            timestamp=value("timestamp"),
        )

    def _rebuild_visible_sheets(self) -> None:
        """Projects hidden checkpoint records into clean human-facing sheets."""
        if self._summary.max_row > 1:
            self._summary.delete_rows(2, self._summary.max_row - 1)

        detail_names = {spec.sheet for spec in _ITEM_SHEETS.values()}
        detail_names.add(_RETRIEVAL_DOCUMENTS_SHEET)
        for name in tuple(detail_names):
            if name in self._workbook.sheetnames:
                self._workbook.remove(self._workbook[name])
        self._item_sheets.clear()

        retrieval_seen: set[tuple[Any, ...]] = set()
        for row_number in range(2, self._checkpoint_sheet.max_row + 1):
            record = self._record_from_checkpoint_row(
                self._checkpoint_sheet[row_number]
            )
            self._summary.append(self._summary_values(record))
            self._append_item_rows(record)
            retrieval_key = (
                record.run_name,
                record.dataset_name,
                record.case_id,
                record.case_fingerprint,
            )
            if retrieval_key not in retrieval_seen and self._append_retrieval_rows(
                record
            ):
                retrieval_seen.add(retrieval_key)

    def _append_item_rows(self, record: EvaluationRecord) -> None:
        """Flattens one record's structured items onto its metric-detail sheet."""
        spec = _ITEM_SHEETS.get(record.metric)
        if spec is None or not record.details:
            return
        items = record.details.get(spec.list_key)
        if not isinstance(items, list) or not items:
            return
        sheet = self._item_sheet(spec)
        for item in items:
            if not isinstance(item, dict):
                continue
            row = [
                record.run_name,
                record.dataset_name,
                record.case_id,
                record.metric,
            ]
            row.extend(item.get(field) for _, field in spec.columns)
            sheet.append(row)

    def _append_retrieval_rows(self, record: EvaluationRecord) -> bool:
        """Writes shared retrieved-document rows once per case; returns whether it did.

        Retrieval records are recognized structurally by a non-empty ``documents``
        list of dicts with a ``rank`` (a shape unique to retrieval metrics), so
        all retrieval metrics map to one shared sheet without duplicate rows or
        a metric column.
        """
        details = record.details
        if not isinstance(details, dict):
            return False
        documents = details.get("documents")
        if not isinstance(documents, list) or not documents:
            return False
        if not all(isinstance(doc, dict) and "rank" in doc for doc in documents):
            return False
        sheet = self._item_sheets.get(_RETRIEVAL_DOCUMENTS_SHEET)
        if sheet is None:
            headers = list(_RETRIEVAL_IDENTITY_COLUMNS) + [
                column for column, _ in _RETRIEVAL_DOCUMENT_COLUMNS
            ]
            sheet = self._workbook.create_sheet(_RETRIEVAL_DOCUMENTS_SHEET)
            self._init_sheet(sheet, headers)
            self._item_sheets[_RETRIEVAL_DOCUMENTS_SHEET] = sheet
        for doc in documents:
            row = [
                record.run_name,
                record.dataset_name,
                record.case_id,
            ]
            row.extend(doc.get(field) for _, field in _RETRIEVAL_DOCUMENT_COLUMNS)
            sheet.append(row)
        return True

    def _item_sheet(self, spec: _ItemSheet) -> Any:
        """Returns the detail sheet for ``spec``, creating it on first use."""
        sheet = self._item_sheets.get(spec.sheet)
        if sheet is None:
            sheet = self._workbook.create_sheet(spec.sheet)
            headers = list(_ITEM_IDENTITY_COLUMNS) + [
                column for column, _ in spec.columns
            ]
            self._init_sheet(sheet, headers)
            self._item_sheets[spec.sheet] = sheet
        return sheet

    @staticmethod
    def _init_sheet(sheet: Any, headers) -> None:
        """Writes a styled header row: bold, frozen, auto-filtered, sized."""
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        headers = list(headers)
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        for index, header in enumerate(headers, start=1):
            width = 48 if header in _WIDE_COLUMNS else max(len(header) + 2, 12)
            sheet.column_dimensions[get_column_letter(index)].width = width


def build_writers(
    output: str | None,
    excel_path: str | None,
    *,
    resume: bool = False,
    report_fields: list[str] | tuple[str, ...] | None = None,
) -> list[EvaluationWriter]:
    """Builds the configured writers for an output mode.

    Args:
        output: One of ``None``, ``"phoenix"``, ``"excel"``, or ``"both"``.
        excel_path: Required when ``output`` includes Excel.
        report_fields: Ordered case fields to project into visible Excel output.

    Returns:
        The list of writers (possibly empty).

    Raises:
        ValueError: For an unknown mode or a missing Excel path.
    """
    if resume and output not in ("excel", "both"):
        raise ValueError(
            "resume=True requires output='excel' or output='both' with an "
            "Excel checkpoint."
        )
    if output is None:
        return []
    if output not in ("phoenix", "excel", "both"):
        raise ValueError(
            f"Unknown output {output!r}; expected 'phoenix', 'excel', 'both', "
            "or None."
        )

    writers: list[EvaluationWriter] = []
    if output in ("phoenix", "both"):
        writers.append(PhoenixEvaluationWriter())
    if output in ("excel", "both"):
        if not excel_path:
            raise ValueError("excel_path is required when output includes Excel.")
        writers.append(
            ExcelEvaluationWriter(
                excel_path, resume=resume, report_fields=report_fields
            )
        )
    return writers
